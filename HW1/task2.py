from data import *
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook

def compute_histogram(filenum, c):
    for idx, dir in enumerate(['Hue', 'Sat', 'Val', 'SV']):
        image_path = f"{dir.lower()}/{filenum}-{idx}.png"

        # Read the image in grayscale
        img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
        
        # Calculate histogram
        hist = cv2.calcHist([img], [0], None, [256], [0, 256])
        
        # Convert histogram to integers and print the array
        hist_int = np.int0(hist.flatten())
        print(hist_int)

        wb = load_workbook("Hist_HSV.xlsx")
        sheet = wb[f"Hist_{dir}"]
        
        if dir == 'SV':
            # Write histogram values to Excel from cell C3 to C258
            for i in range(len(hist_int)):
                cell = sheet.cell(row=i+3, column=c)  # Starting from C3 (row index 3, column index 3)
                cell.value = hist_int[i]

        # Save the workbook
        wb.save("Hist_HSV.xlsx")
        wb.close()


def main():
    for idx, file in enumerate(files):
        print(idx)
        compute_histogram(file,3 + idx)

main()
