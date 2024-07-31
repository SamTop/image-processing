import cv2
import numpy as np
from collections import defaultdict
from data import files

colors = None
# Load the image
for filename in files:
    image_path = f"face/{filename}.png"  # Replace with your image path
    image = cv2.imread(image_path)

    if image is None:
        raise ValueError("Image not found at the specified path.")

    # Flatten the image array to a list of colors
    temp = image.reshape((-1, 3))
    if colors is None:
        colors = np.empty((0, temp.shape[1]), dtype=temp.dtype)
    colors = np.concatenate((colors, temp), axis=0)

# Filter out black colors (0, 0, 0) and get unique colors
unique_colors = np.unique(colors[colors.sum(axis=1) != 0], axis=0)

# Initialize dictionaries to store the results
count = defaultdict(int)
min_val = {i: float('inf') for i in range(256)}
max_val = {i: float('-inf') for i in range(256)}
mean = defaultdict(float)
mean2 = defaultdict(float)

# Temporary storage to compute means
sum_rb = defaultdict(float)
sum_rb_squared = defaultdict(float)

# Compute the required quantities
for color in unique_colors:
    R, G, B = color
    rb_mean = (R + B) / 2
    rb_mean_squared = rb_mean ** 2
    
    # Update the count
    count[G] += 1
    
    # Update the min and max values
    if rb_mean < min_val[G]:
        min_val[G] = rb_mean
    if rb_mean > max_val[G]:
        max_val[G] = rb_mean
    
    # Update sums for mean and mean squared
    sum_rb[G] += rb_mean
    sum_rb_squared[G] += rb_mean_squared

# Calculate means from the sums
for G in range(256):
    if count[G] > 0:
        mean[G] = sum_rb[G] / count[G]
        mean2[G] = sum_rb_squared[G] / count[G]
    else:
        # If no colors were found for a particular G, set these values to None or appropriate defaults
        min_val[G] = None
        max_val[G] = None
        mean[G] = None
        mean2[G] = None

import openpyxl

file_path = 'RGB.xlsx'  # Replace with your Excel file path
workbook = openpyxl.load_workbook(file_path)

# Select the worksheet you want to work with
worksheet_name = 'RB(G)11'  # Replace with your worksheet name
worksheet = workbook.worksheets[1]

# # Output results
# for G in range(256):
#     if count[G] > 0:
#         print(f'Green: {G}')
#         print(f'Count: {count[G]}')
#         print(f'Min (R+B)/2: {min_val[G]}')
#         print(f'Max (R+B)/2: {max_val[G]}')
#         print(f'Mean (R+B)/2: {mean[G]}')
#         print(f'Mean ((R+B)/2)^2: {mean2[G]}')
#         print('---')

# Load the Excel file

# Specify the cell range and the number to input
cell_range = 'C3:C258'  # Replace with your desired cell range

# Split the cell range to get the start and end cells
start_cell, end_cell = cell_range.split(':')

# Get the column and row indices for the start and end cells
start_row, col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
end_row, _ = openpyxl.utils.cell.coordinate_to_tuple(end_cell)

# Iterate over the specified cell range and input the number
for row in range(start_row, end_row + 1):
    G = row - start_row
    if count[G] > 0:
        worksheet.cell(row=row, column=col).value = count[G]
        worksheet.cell(row=row, column=col+1).value = min_val[G]
        worksheet.cell(row=row, column=col+2).value = max_val[G]
        worksheet.cell(row=row, column=col+3).value = mean[G]
        worksheet.cell(row=row, column=col+4).value = mean2[G]
    else:
        worksheet.cell(row=row, column=col).value = 0
        worksheet.cell(row=row, column=col+1).value = 0
        worksheet.cell(row=row, column=col+2).value = 0
        worksheet.cell(row=row, column=col+3).value = 0
        worksheet.cell(row=row, column=col+4).value = 0

# Save the changes to the Excel file
workbook.save(file_path)
